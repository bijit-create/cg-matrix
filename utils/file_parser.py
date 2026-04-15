"""Parse uploaded files: PDF, DOCX, Excel."""

import io

def parse_pdf(file_bytes: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    return "\n".join(page.extract_text() or "" for page in reader.pages)

def parse_docx(file_bytes: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

def parse_excel(file_bytes: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"--- Sheet: {sheet} ---")
        for row in ws.iter_rows(values_only=True):
            lines.append(" | ".join(str(c or "") for c in row))
    return "\n".join(lines)

def parse_file(filename: str, file_bytes: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "pdf": return parse_pdf(file_bytes)
    if ext == "docx": return parse_docx(file_bytes)
    if ext in ("xlsx", "xls"): return parse_excel(file_bytes)
    raise ValueError(f"Unsupported file type: .{ext}")
