"""Export questions to Excel + ZIP."""

import io
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_excel(questions: list, metadata: dict) -> bytes:
    wb = openpyxl.Workbook()

    # --- Questions sheet ---
    ws = wb.active
    ws.title = "Questions"

    headers = ["ID", "Cell", "Type", "Question Stem", "Correct Answer", "Rationale",
               "Needs Image", "Option A", "Option A Correct", "Option A Why Wrong",
               "Option B", "Option B Correct", "Option B Why Wrong",
               "Option C", "Option C Correct", "Option C Why Wrong",
               "Option D", "Option D Correct", "Option D Why Wrong",
               "Step 1", "Step 1 Correct", "Step 1 Fix",
               "Step 2", "Step 2 Correct", "Step 2 Fix",
               "Step 3", "Step 3 Correct", "Step 3 Fix",
               "Step 4", "Step 4 Correct", "Step 4 Fix",
               "Match Pairs", "Arrange Items"]

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for ri, q in enumerate(questions, 2):
        ws.cell(row=ri, column=1, value=q.get("id", ""))
        ws.cell(row=ri, column=2, value=q.get("cell", ""))
        ws.cell(row=ri, column=3, value=q.get("type", "mcq"))
        ws.cell(row=ri, column=4, value=q.get("stem", ""))
        ws.cell(row=ri, column=5, value=q.get("answer", ""))
        ws.cell(row=ri, column=6, value=q.get("rationale", ""))
        ws.cell(row=ri, column=7, value="Yes" if q.get("needs_image") else "No")

        # Options
        for oi, opt in enumerate(q.get("options", [])[:4]):
            base_col = 8 + oi * 3
            ws.cell(row=ri, column=base_col, value=opt.get("text", ""))
            ws.cell(row=ri, column=base_col + 1, value="Yes" if opt.get("correct") else "No")
            ws.cell(row=ri, column=base_col + 2, value=opt.get("why_wrong", ""))

        # Steps
        for si, step in enumerate(q.get("steps", [])[:4]):
            base_col = 20 + si * 3
            ws.cell(row=ri, column=base_col, value=step.get("text", ""))
            ws.cell(row=ri, column=base_col + 1, value="Correct" if step.get("correct") else "Incorrect")
            ws.cell(row=ri, column=base_col + 2, value=step.get("fix", ""))

        # Pairs & Items
        pairs = q.get("pairs", [])
        ws.cell(row=ri, column=32, value=" | ".join(pairs) if pairs else "")
        items = q.get("items", [])
        ws.cell(row=ri, column=33, value=" → ".join(items) if items else "")

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Learning Objective", metadata.get("lo", "")),
        ("Skill", metadata.get("skill", "")),
        ("Grade", metadata.get("grade", "")),
        ("Subject", metadata.get("subject", "")),
        ("Total Questions", len(questions)),
        ("Types", ", ".join(set(q.get("type", "mcq") for q in questions))),
        ("Export Date", datetime.now().isoformat()),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=str(v))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
